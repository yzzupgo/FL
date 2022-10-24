import datetime
import json
# import ijson
import os
import math
import numpy as np
import socket
import util
import openpyxl
import random
import copy
import pandas as pd

from whf import Get_sus
from whf import *
import mbfl.mbfl_for
import mbfl.command
import sbfl.sbfl_for
import sbfl.command
from CHMBFL_Flow import Cluster_Fom
from ChmbflSus import Tools
from SirSus import SirTools
from SirSus import data_dirpath
from mbfl.mutpolyn import mutation_trick
from mbfl.mutpolyn import new_mutation_trick
from data_codeflaws import MutationRule


mbfl_for_list = [
    "mbfl.mbfl_for.Tarantula",
    "mbfl.mbfl_for.Op2",
    "mbfl.mbfl_for.Jaccard",
    "mbfl.mbfl_for.Ochiai",
    "mbfl.mbfl_for.Dstar",
    "mbfl.mbfl_for.GP13",
    "mbfl.mbfl_for.Naish1",
    "mbfl.mbfl_for.Barinel",
    "mbfl.mbfl_for.muse",
    "mbfl.mbfl_for.nsus",
]

sbfl_for_list = [
    "sbfl.sbfl_for.Tarantula",
    "sbfl.sbfl_for.Op2",
    "sbfl.sbfl_for.Jaccard",
    "sbfl.sbfl_for.Ochiai",
    "sbfl.sbfl_for.Dstar",
    "sbfl.sbfl_for.GP13",
    "sbfl.sbfl_for.Naish1",
    "sbfl.sbfl_for.Barinel",
]

class Timecreat:
    def __init__(self, totalFile):
        self.sTime = datetime.datetime.now()

        self.fileData = {
            'name': '',
            'done': -1,
            'total': totalFile,
            'rest': 0,
        }

        self.methodData = {
            'sTime': 0,
            'done': 0,
            'total': 0,
        }

    def newFile(self, fileName, methodsNum):
        nTime = datetime.datetime.now()
        costTime = nTime-self.sTime
        self.fileData['done'] += 1
        self.fileData['rest'] = costTime/self.fileData['done']*(self.fileData['total']-self.fileData['done']) \
            if self.fileData['done'] else '???'
        self.fileData['name'] = fileName

        print("{},    {}/{}, {}, 以用时:{}, 剩余时间:{}".format(
            nTime,
            self.fileData['done'],
            self.fileData['total'],
            self.fileData['name'],
            costTime,
            self.fileData['rest'],
            ))

        self.methodData = {
            'name': '',
            'sTime': datetime.datetime.now(),
            'lTime': datetime.datetime.now(),
            'done': 0,
            'total': methodsNum,
        }

    def newMethod(self, methodName):
        self.methodData['lTime'] = datetime.datetime.now()
        self.methodData['name'] = methodName

        print("{},    {}/{}  {}-rest{}    {}/{}-{}".format(
            self.methodData['lTime'],
            self.fileData['done'],
            self.fileData['total'],
            self.fileData['name'],
            self.fileData['rest'],
            self.methodData['done'],
            self.methodData['total'],
            self.methodData['name'],
        ))

    def doneMethod(self):
        nTime = datetime.datetime.now()
        costTime = nTime - self.methodData['lTime']
        restTime = costTime/self.methodData['done']*(self.methodData['total']-self.methodData['done']) \
            if self.methodData['done'] else '???'
        self.methodData['lTime'] = nTime
        self.methodData['done'] += 1

        print("{},         cost:{}    rest:{}".format(
            nTime,
            costTime,
            restTime,
        ))

class WB:
    def __init__(self, path):
        self.wb = openpyxl.load_workbook(path)


class Data(object):
    def __init__(self, path):
        self._data = 'empty'
        self.path = path

    # 访问器 getter方法
    @property
    def data(self):
        try:
            if self._data == 'empty':
                self._data = SirTools().sir_rules(self.path, covread=True, firstread=True, oneline=True)
        except Exception as e:
            self._data = False
            print(self.path, 'less', e)
        return self._data





class Sirsus:
    def __init__(self):
        self.continuecreat = True
        # self.continuecreat = False
        # self.times()
        # self.timestest()
        self.yojan()
        # self.newStatic()
        # self.reWriteDif()


    # 不需要倍数，需要重复次数
    def times(self):
        clasterfunction = [
            ['Sbfl', Get_sus().Sbfl],
            ['Mbfl', Get_sus().Mbfl],
            ['Mcbfl', Get_sus().Mcbfl],

            ['Last2First', Get_sus().Last2First],
            ['DifferentOperator', Get_sus().DifferentOperator],
            # ['RandomMix', Get_sus().RandomMix],
            ['Random', Get_sus().Random],

            # ['ED.SBFL', Get_sus().ErrorDistribution, [False]],
            ['ED.MBFL', Get_sus().ErrorDistribution, [True]],

            ['NS.RANDOM', Get_sus().NS, [0]],
            ['NS.SBFL', Get_sus().NS, [2]],
            ['NS.MBFL', Get_sus().NS, [1]],

            ['MutCluster.kmeans', Get_sus().MutCluster, [False, True]],
            # ['MutCluster.kmeans.in', Get_sus().MutCluster, [False, False]],
            ['MutCluster', Get_sus().MutCluster, [True, True]],
            ['MutCluster.in', Get_sus().MutCluster, [True, False]],

            ['NS.mbfl^MC', Get_sus().NSpMC],
            ['NS.mbfl^ED', Get_sus().NSpED],
            ['MC.mseer^ED', Get_sus().MCpED],
            ['ED^NS^MC.mseer', Get_sus().EDpNSpMC],
        ]

        ftitle = [
            'Method',
            'times',
            'repeat',
            'Statement',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        repeattimes = 5
        timerange = list(range(1, 1+1))

        for i in range(15):
            ftitle.append('Operator Type %s' % str(i))
        ftitle += [
            'None Accurate',
            'Pare Accurate',
            'Accurate',
        ]
        maxfaultnum = 5

        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile-sir/K')
        if not os.path.exists(docpath):
            os.makedirs(docpath)

        versionnum = len(SirTools().version_path())
        st = datetime.datetime.now()

        for file_i, [project, faultnum, version] in enumerate(SirTools().version_path()):

            nt = datetime.datetime.now()
            file_resttime = (nt-st)/file_i*(versionnum-file_i) if file_i > 0 else '???'
            print("{},    {}/{}, {}-{}-{}, 以用时:{}".format(
                nt,
                file_i,
                versionnum,
                project,
                faultnum,
                version,
                nt-st)
            )

            file = '%s-%s-%s' % (project, faultnum, version)
            path = os.path.join(data_dirpath, project, faultnum, version)
            # filepath_undo = os.path.join(docpath, '%s_%s-undo.xlsx' % (file_i, file))
            # filepath_do = os.path.join(docpath, '%s_%s.xlsx' % (file_i, file))
            # filepath_undo = os.path.join(docpath, '%s-undo.xlsx' % (file))
            filepath_do = os.path.join(docpath, '%s.xlsx' % (file))

            # 表格初始化
            passlist = []
            wb = False
            if self.continuecreat:
                # if os.path.exists(filepath_do):
                #     continue
                if os.path.exists(filepath_do):
                    df = pd.read_excel(filepath_do, sheet_name='exam-best', header=[1])
                    for i in df.index.values:
                        row_data_read = df.loc[i, df.columns.values].to_dict()
                        passlist.append([
                            row_data_read['Method'],
                            row_data_read['times'],
                            row_data_read['repeat'],
                            row_data_read['Statement'],
                        ])
                    wb = openpyxl.load_workbook(filepath_do)
            if not wb:
                wb = openpyxl.Workbook()
                del wb['Sheet']
                for ws_name in ['exam-best', 'exam-average', 'exam-worst']:
                    ws = wb.create_sheet(ws_name)
                    ws.cell(1, 1, file)
                    for j, title in enumerate(ftitle):
                        ws.cell(2, j+1, title)
                    for j, mbfl_for in enumerate(mbfl_for_list):
                        ws.cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                        for k in range(maxfaultnum):
                            ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                            ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+maxfaultnum+k, 'rank%s' % str(k+1))

            sir_data = Data(path)

            nnum = len(passlist)
            tnum = len(clasterfunction)*len(timerange)*repeattimes*3
            lt = datetime.datetime.now()
            for parameter in clasterfunction:
                dosave = False
                for times in timerange:
                    for repeat in range(1, repeattimes+1):
                        method = parameter[0]
                        if [method, times, repeat, 'frequency'] in passlist:
                            continue
                        dosave = False
                        function = parameter[1]

                        if not sir_data.data:
                            continue
                        dosave = True

                        print("{},    {}/{}-filerest:{},    {}/{}-{}/{}/{}".format(
                            datetime.datetime.now(),
                            file_i,
                            versionnum,
                            file_resttime,
                            nnum,
                            tnum,
                            method,
                            times,
                            repeat,
                        ))
                        lt = datetime.datetime.now()

                        if parameter[0] in ['Sbfl', 'Mbfl', 'Mcbfl']:
                            sus_data_json = function(sir_data.data)
                        else:
                            if len(parameter) <= 2:
                                sus_data_json = function(sir_data.data, times=times)
                            else:
                                sus_data_json = function(sir_data.data, times=times, seting=parameter[2])


                        for tie in ['best', 'average', 'worst']:
                            ws_name = 'exam-%s' % tie
                            ws = wb[ws_name]
                            for numi, word in enumerate(['max', 'ave', 'frequency']):

                                sheetdata_list = [
                                    method,
                                    times,
                                    repeat,
                                    word,
                                ]
                                if not sus_data_json:
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        ws.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                    continue
                                if word not in sus_data_json:
                                    wordMessage = sus_data_json
                                else:
                                    wordMessage = sus_data_json[word]

                                sheetdata_list += [
                                    len(sus_data_json['Fault_Record']),
                                    sus_data_json['totaltime'],
                                    sus_data_json['fomnum'],
                                    sus_data_json['homnum'],
                                    '',
                                ]

                                if 'variety' in sus_data_json:
                                    varietys = sus_data_json['variety']
                                    sum_varietys = sum(varietys)
                                    sheetdata_list += list(map(lambda x: 0 if sum_varietys == 0 else x/sum_varietys,
                                                               varietys))
                                else:
                                    sheetdata_list += [0 for i in range(15)]

                                if 'precision' in sus_data_json:
                                    precisions = sus_data_json['precision']
                                    sum_precisions = sum(precisions)
                                    sheetdata_list += list(map(lambda x: 0 if sum_precisions == 0 else x/sum_precisions,
                                                               precisions))
                                else:
                                    sheetdata_list += [0 for i in range(3)]

                                for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):

                                    if parameter[0] == 'Sbfl':
                                        formula = mbfl_for.replace('mbfl', 'sbfl')
                                    else:
                                        formula = mbfl_for

                                    examlist = []
                                    ranklist = []
                                    for fault_ith in range(maxfaultnum):
                                        if formula not in wordMessage:
                                            examlist.append('')
                                            ranklist.append('')
                                            continue
                                        if fault_ith > len(wordMessage[formula]['rank_%s' % tie])-1:
                                            examlist.append('')
                                            ranklist.append('')
                                        else:
                                            examlist.append(wordMessage[formula]['exam_%s' % tie][fault_ith])
                                            ranklist.append(wordMessage[formula]['rank_%s' % tie][fault_ith])
                                    sheetdata_list += examlist
                                    sheetdata_list += ranklist
                                for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                    ws.cell(nnum + numi + 3, sheetdata_i+1, sheetdata)
                        nnum += 3

                        nt = datetime.datetime.now()
                        print("{},    cost:{}".format(
                            datetime.datetime.now(),
                            '???' if nnum == 0 else nt-lt,
                        ))


                # 存储文件
                if dosave:
                    wb.save(filepath_do)
                    print('文件保存 %s ' % filepath_do)
                # return

            # os.remove(filepath_undo)
            # wb.save(filepath_do)
            # print('文件保存 %s ' % filepath_do)
        return


    # 不需要倍数，需要重复次数
    def newStatic(self):

        ftitle = [
            '版本',
            '代码行数',
            '测试用例个数',
            '错误数量',
            '生成一阶变异体数量',
        ]

        wb = openpyxl.Workbook()
        ws = wb['Sheet']
        for j, title in enumerate(ftitle):
            ws.cell(1, j+1, title)

        line = 2

        for file_i, [project, faultnum, version] in enumerate(SirTools().version_path()):

            path = os.path.join(data_dirpath, project, faultnum, version)
            name = '_'.join(path.split('/')[3:])
            sir_data = SirTools().sir_rules(path, covread=True, firstread=True, oneline=True)
            if not sir_data:
                continue
            writeData = [
                '-'.join([project, faultnum, version]),
                sir_data[name]['linelen'],
                len(sir_data[name]['or_list']),
                len(sir_data[name]['Fault_Record']),
                len(sir_data[name]['fom_list']),
            ]
            for j, value in enumerate(writeData):
                ws.cell(line, j+1, value)
            line += 1
        wb.save('需要统计的数据集信息.xlsx')



        return


    # 不需要倍数，需要重复次数
    def timestest(self):
        clasterfunction = [
            ['Sbfl', Get_sus().Sbfl],
            ['Mbfl', Get_sus().Mbfl],
            ['Mcbfl', Get_sus().Mcbfl],

            ['Last2First', Get_sus().Last2First],
            ['DifferentOperator', Get_sus().DifferentOperator],
            # ['RandomMix', Get_sus().RandomMix],
            ['Random', Get_sus().Random],

            # ['ED.SBFL', Get_sus().ErrorDistribution, [False]],
            ['ED.MBFL', Get_sus().ErrorDistribution, [True]],

            ['NS.RANDOM', Get_sus().NS, [0]],
            ['NS.SBFL', Get_sus().NS, [2]],
            ['NS.MBFL', Get_sus().NS, [1]],

            ['MutCluster.kmeans', Get_sus().MutCluster, [False, True]],
            # ['MutCluster.kmeans.in', Get_sus().MutCluster, [False, False]],
            ['MutCluster', Get_sus().MutCluster, [True, True]],
            ['MutCluster.in', Get_sus().MutCluster, [True, False]],

            ['NS.mbfl^MC', Get_sus().NSpMC],
            ['NS.mbfl^ED', Get_sus().NSpED],
            ['MC.mseer^ED', Get_sus().MCpED],
            ['ED^NS^MC.mseer', Get_sus().EDpNSpMC],
        ]


        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile-sir/K')
        if not os.path.exists(docpath):
            os.makedirs(docpath)


        for filename in os.listdir(docpath):
            filepath_do = os.path.join(docpath, filename)

            # 表格初始化
            passlist = []
            df = pd.read_excel(filepath_do, sheet_name='exam-best', header=[1])
            for i in df.index.values:
                row_data_read = df.loc[i, df.columns.values].to_dict()
                passlist.append(row_data_read['Method'])

            less = []
            for parameter in clasterfunction:
                method = parameter[0]
                if method not in passlist:
                    less.append(method)
            if len(less) > 0:
                print(filepath_do)
                print(less)
        return

    # 不需要倍数，需要重复次数
    def yojan(self):
        reduction_list = [
            'testanalysis',
            'samping',
            'some',
            'wsome',
            # 'fomanalysis',
        ]

        clasterfunction = [
            # ['Last2First', Get_sus().Last2First],
            # ['DifferentOperator', Get_sus().DifferentOperator],
            # ['Random', Get_sus().Random],

            # ['ED.SBFL', Get_sus().ErrorDistribution, [False]],
            ['ED.MBFL', Get_sus().ErrorDistribution, [True]],

            ['NS.RANDOM', Get_sus().NS, [0]],
            ['NS.SBFL', Get_sus().NS, [2]],
            # ['NS.MBFL', Get_sus().NS, [1]],

            # ['MutCluster.kmeans', Get_sus().MutCluster, [False, True]],
            # ['MutCluster.kmeans.in', Get_sus().MutCluster, [False, False]],
            ['MutCluster', Get_sus().MutCluster, [True, True]],
            # ['MutCluster.in', Get_sus().MutCluster, [True, False]],

            # ['NS.mbfl^MC', Get_sus().NSpMC],
            # ['NS.mbfl^ED', Get_sus().NSpED],
            # ['MC.mseer^ED', Get_sus().MCpED],
            # ['ED^NS^MC.mseer', Get_sus().EDpNSpMC],
        ]

        # 创建ftitle
        ftitle = [
            'Method',
            'Reduction',
            'times',
            'repeat',
            'Statement',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        for i in range(15):
            ftitle.append('Operator Type %s' % str(i))
        ftitle += ['None Accurate', 'Pare Accurate', 'Accurate', ]


        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile-sir/yojan-need')
        if not os.path.exists(docpath):
            os.makedirs(docpath)
        maxfaultnum = 5
        repeattimes = 5
        timeslist = [1.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]

        timeCreat = Timecreat(len(SirTools().version_path()))

        for file_i, [project, faultnum, version] in enumerate(SirTools().version_path()):

            # 数据初始化
            file = '%s-%s-%s' % (project, faultnum, version)
            path = os.path.join(data_dirpath, project, faultnum, version)
            # filepath_undo = os.path.join(docpath, '%s-undo.xlsx' % (file))
            filepath_do = os.path.join(docpath, '%s.xlsx' % (file))
            methodsNum = len(clasterfunction)*len(reduction_list)*len(timeslist)
            timeCreat.newFile(file, methodsNum)

            # 断点续传
            wb = False
            passlist = []
            if self.continuecreat:
                if os.path.exists(filepath_do):
                    df = pd.read_excel(filepath_do, sheet_name='exam-best', header=[1])
                    for i in df.index.values:
                        row_data_read = df.loc[i, df.columns.values].to_dict()
                        if not row_data_read['Statement'] == 'max':
                            continue
                        if not row_data_read['Reduction'] == 'testanalysis':
                            continue
                        passlist.append([
                            row_data_read['Method'],
                            row_data_read['times'],
                            row_data_read['repeat'],
                        ])
                    wb = openpyxl.load_workbook(filepath_do)
                    for ws_name in ['exam-best', 'exam-average', 'exam-worst']:
                        ws = wb[ws_name]
                        ws.nowLine = len(passlist)*len(reduction_list)*3+3

            # 表格初始化
            if not wb:
                wb = openpyxl.Workbook()
                del wb['Sheet']
                for ws_name in ['exam-best', 'exam-average', 'exam-worst']:
                    ws = wb.create_sheet(ws_name)
                    ws.cell(1, 1, file)
                    for j, title in enumerate(ftitle):
                        ws.cell(2, j+1, title)
                    for j, mbfl_for in enumerate(mbfl_for_list):
                        ws.cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                        for k in range(maxfaultnum):
                            ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                            ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+maxfaultnum+k, 'rank%s' % str(k+1))
                    ws.nowLine = 3

            # 数据读取
            try:
                sir_data = SirTools().sir_rules(path, covread=True, firstread=True, oneline=True)
            except Exception as e:
                print(path, 'less', e)
                continue

            for parameter in clasterfunction:
                for times in timeslist:
                    for repeat in range(1, repeattimes+1):
                        method = parameter[0]
                        function = parameter[1]

                        timeCreat.newMethod('%s-%s-%s'%(method, times, repeat))

                        # 断点续传
                        if [method, times, repeat] in passlist:
                            continue

                        reducefunction = list(map(lambda key:eval('Reduction(%s).%s' % (str(times), key)),
                                                  reduction_list))
                        if len(parameter) <= 2:
                            sus_data_json = function(sir_data, times=1, reducefunction=reducefunction)
                        else:
                            sus_data_json = function(sir_data, times=1, seting=parameter[2], reducefunction=reducefunction)

                        for tie in ['best', 'average', 'worst']:
                            ws_name = 'exam-%s' % tie
                            ws = wb[ws_name]
                            for reductionKey in reduction_list:
                                for word in ['max', 'ave', 'frequency']:
                                    sheetdata_list = [
                                        method,
                                        reductionKey,
                                        times,
                                        repeat,
                                        word,
                                    ]
                                    if not sus_data_json:
                                        for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                            ws.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                        continue
                                    susData = sus_data_json[reductionKey]
                                    sheetdata_list += [
                                        len(susData['Fault_Record']),
                                        susData['totaltime'],
                                        susData['fomnum'],
                                        susData['homnum'],
                                        '',
                                    ]

                                    varietys = susData['variety']
                                    sum_varietys = sum(varietys)
                                    sheetdata_list += list(map(lambda x: 0 if sum_varietys == 0 else x/sum_varietys,
                                                               varietys))

                                    precisions = susData['precision']
                                    sum_precisions = sum(precisions)
                                    sheetdata_list += list(map(lambda x: 0 if sum_precisions == 0 else x/sum_precisions,
                                                               precisions))

                                    for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                        examlist = []
                                        ranklist = []
                                        for fault_ith in range(maxfaultnum):
                                            if fault_ith > len(susData[word][mbfl_for]['rank_%s' % tie])-1:
                                                examlist.append('')
                                                ranklist.append('')
                                            else:
                                                examlist.append(susData[word][mbfl_for]['exam_%s' % tie][fault_ith])
                                                ranklist.append(susData[word][mbfl_for]['rank_%s' % tie][fault_ith])
                                        sheetdata_list += examlist
                                        sheetdata_list += ranklist
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        ws.cell(ws.nowLine, sheetdata_i+1, sheetdata)
                                    ws.nowLine += 1
                        timeCreat.doneMethod()

                    # 存储文件
                    wb.save(filepath_do)
                    print('文件保存 %s ' % filepath_do)
                    # return

            # os.remove(filepath_undo)
            wb.save(filepath_do)
            print('文件保存 %s ' % filepath_do)
        return


    # 不需要倍数，需要重复次数
    def reWriteDif(self):
        clasterfunction = [
            ['DifferentOperator', Get_sus().DifferentOperator],
        ]

        ftitle = [
            'Method',
            'times',
            'repeat',
            'Statement',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]

        for i in range(15):
            ftitle.append('Operator Type %s' % str(i))
        ftitle += [
            'None Accurate',
            'Pare Accurate',
            'Accurate',
        ]
        maxfaultnum = 5
        repeattimes = 5

        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile-sir/K')
        if not os.path.exists(docpath):
            os.makedirs(docpath)

        for file_i, file in enumerate(os.listdir(docpath)):
            project, faultnum, version = file.split('.')[0].split('-')

            print("{},    {}/{}, {}-{}-{}".format(
                datetime.datetime.now(),
                file_i,
                len(os.listdir(docpath)),
                project,
                faultnum,
                version,)
            )

            path = os.path.join(data_dirpath, project, faultnum, version)
            filepath_do = os.path.join(docpath, file)

            # 数据检测
            df = pd.read_excel(filepath_do, sheet_name='exam-best', header=[1])
            reCreat = []
            for i in df.index.values:
                row_data_read = df.loc[i, df.columns.values].to_dict()
                if not 'DifferentOperator' == row_data_read['Method']:
                    continue
                if not row_data_read['Homnum'] == 0:
                    continue
                reCreat.append(i)
            if len(reCreat) == 0:
                continue
            print(filepath_do, 'diferror')
            continue

            wb = openpyxl.load_workbook(filepath_do)
            # a = wb['exam-best'].cell(1, 1)

            # 数据读取
            try:
                sir_data = SirTools().sir_rules(path, covread=True, firstread=True, oneline=True)
            except Exception as e:
                print(path, 'less', e)
                continue

            times = 1
            parameter = ['DifferentOperator', Get_sus().DifferentOperator]
            for repeat in range(1, repeattimes+1):
                method = parameter[0]
                function = parameter[1]

                print("{},    {}/{}-{}".format(
                    datetime.datetime.now(),
                    file_i,
                    len(os.listdir(docpath)),
                    repeat,
                ))

                if len(parameter) <= 2:
                    sus_data_json = function(sir_data, times=times)
                else:
                    sus_data_json = function(sir_data, times=times, seting=parameter[2])

                for tie in ['best', 'average', 'worst']:
                    ws_name = 'exam-%s' % tie
                    ws = wb[ws_name]
                    for numi, word in enumerate(['max', 'ave', 'frequency']):
                        line = reCreat[(repeat-1)*3 + numi]+3
                        sheetdata_list = [
                            method,
                            times,
                            repeat,
                            word,
                        ]
                        if not sus_data_json:
                            for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                ws.cell(file_i + 3, sheetdata_i+1, sheetdata)
                            continue
                        sheetdata_list += [
                            len(sus_data_json['Fault_Record']),
                            sus_data_json['totaltime'],
                            sus_data_json['fomnum'],
                            sus_data_json['homnum'],
                            '',
                        ]

                        varietys = sus_data_json['variety']
                        sum_varietys = sum(varietys)
                        sheetdata_list += list(map(lambda x: 0 if sum_varietys == 0 else x/sum_varietys,
                                                   varietys))

                        precisions = sus_data_json['precision']
                        sum_precisions = sum(precisions)
                        sheetdata_list += list(map(lambda x: 0 if sum_precisions == 0 else x/sum_precisions,
                                                   precisions))

                        for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                            examlist = []
                            ranklist = []
                            for fault_ith in range(maxfaultnum):
                                if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % tie])-1:
                                    examlist.append('')
                                    ranklist.append('')
                                else:
                                    examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % tie][fault_ith])
                                    ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % tie][fault_ith])
                            sheetdata_list += examlist
                            sheetdata_list += ranklist
                        for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                            ws.cell(line, sheetdata_i+1, sheetdata)

            wb.save(filepath_do)
            print('文件保存 %s ' % filepath_do)
        return

def test():
    project, faultnum, version = 'printtokens2', 'four_faults', 'v1'
    path = os.path.join(data_dirpath, project, faultnum, version)
    data = SirTools().sir_rules(path, covread=True, firstread=True, oneline=True)
    name = '_'.join(path.split('/')[3:])
    print(len(data[name]['fom_list']))
    print(len(data[name]['hom_out_list']))
    print(len(data[name]['oneline_out_list']))



if __name__ == '__main__':
    Sirsus()
    # test()


